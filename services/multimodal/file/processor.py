"""文件处理器 - PDF/Word/Excel等企业文档解析"""

from shared.utils import get_logger
from services.multimodal.router import BaseProcessor

logger = get_logger(__name__)


class FileProcessor(BaseProcessor):
    """文件处理器 - 支持多种企业文档格式"""

    SUPPORTED_EXTENSIONS = {"pdf", "docx", "xlsx", "csv", "txt"}

    async def process(self, content: bytes | str, metadata: dict) -> dict:
        """解析文件内容，返回结构化文本"""
        filename = metadata.get("filename", "")
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

        if isinstance(content, str):
            return {"text": content, "format": "text", "pages": 1}

        try:
            if ext == "pdf":
                return await self._parse_pdf(content)
            elif ext == "docx":
                return await self._parse_docx(content)
            elif ext == "xlsx":
                return await self._parse_xlsx(content)
            elif ext == "csv":
                return await self._parse_csv(content)
            elif ext == "txt":
                text = content.decode("utf-8", errors="replace")
                return {"text": text, "format": "txt", "pages": 1}
            else:
                return {"text": "", "format": ext, "error": f"不支持的文件格式: {ext}"}

        except Exception as e:
            logger.error("File processing failed", filename=filename, error=str(e))
            return {"text": "", "format": ext, "error": str(e)}

    async def _parse_pdf(self, content: bytes) -> dict:
        """解析PDF文件"""
        import io
        from PyPDF2 import PdfReader

        reader = PdfReader(io.BytesIO(content))
        pages_text = []
        for page in reader.pages:
            text = page.extract_text() or ""
            pages_text.append(text)

        full_text = "\n\n".join(pages_text)
        return {"text": full_text, "format": "pdf", "pages": len(reader.pages)}

    async def _parse_docx(self, content: bytes) -> dict:
        """解析Word文档"""
        import io
        from docx import Document

        doc = Document(io.BytesIO(content))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        full_text = "\n".join(paragraphs)
        return {"text": full_text, "format": "docx", "paragraphs": len(paragraphs)}

    async def _parse_xlsx(self, content: bytes) -> dict:
        """解析Excel文件"""
        import io
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True)
        sheets_data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])
            sheets_data[sheet_name] = rows

        # 转为文本表示
        text_parts = []
        for name, rows in sheets_data.items():
            text_parts.append(f"=== Sheet: {name} ===")
            for row in rows[:100]:  # 限制行数
                text_parts.append(" | ".join(row))

        return {"text": "\n".join(text_parts), "format": "xlsx", "sheets": list(sheets_data.keys())}

    async def _parse_csv(self, content: bytes) -> dict:
        """解析CSV文件"""
        text = content.decode("utf-8", errors="replace")
        lines = text.strip().split("\n")
        return {"text": text, "format": "csv", "rows": len(lines)}
